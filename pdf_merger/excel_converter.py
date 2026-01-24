"""
Excel converter module.
Handles converting Excel files to PDF format.
"""

from pathlib import Path
from typing import Optional

from .logger import get_logger
from .enums import EXCEL_FILE_EXTENSIONS

logger = get_logger("excel_converter")


def _safe_str(value) -> str:
    """
    Safely convert a value to string, handling None and other types.
    
    Args:
        value: Value to convert
        
    Returns:
        String representation of the value
    """
    if value is None:
        return ''
    return str(value)


def convert_excel_to_pdf(excel_path: Path, output_path: Path) -> bool:
    """
    Convert an Excel file to PDF format using openpyxl and reportlab.
    
    Args:
        excel_path: Path to the Excel file (.xlsx or .xls)
        output_path: Path where the PDF will be saved
        
    Returns:
        True if successful, False otherwise
        
    Raises:
        ImportError: If required libraries are not installed
    """
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        return False
    
    if excel_path.suffix.lower() not in EXCEL_FILE_EXTENSIONS:
        logger.error(f"File is not an Excel file: {excel_path}")
        return False
    
    try:
        # Import required libraries
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl library is required for Excel file conversion. "
                "Install with: pip install openpyxl"
            )
        
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            raise ImportError(
                "reportlab library is required for PDF generation. "
                "Install with: pip install reportlab"
            )
        
        # Load the Excel workbook
        try:
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel file {excel_path.name}: {e}")
            return False
        
        # Get the active sheet (or first sheet)
        sheet = wb.active
        
        # Collect all data from the sheet
        data = []
        max_col = sheet.max_column
        max_row = sheet.max_row
        
        if max_row == 0 or max_col == 0:
            logger.warning(f"Excel file {excel_path.name} appears to be empty")
            # Create an empty PDF
            doc = SimpleDocTemplate(str(output_path), pagesize=letter)
            doc.build([])
            return True
        
        # Read all rows
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
            # Convert row to list of strings, handling None values
            row_data = [_safe_str(cell) for cell in row]
            data.append(row_data)
        
        # Create PDF document
        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        
        # Create table from data
        table = Table(data)
        
        # Style the table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ])
        
        table.setStyle(style)
        
        # Build PDF
        elements = [table]
        doc.build(elements)
        
        # Verify the output file was created
        if not output_path.exists():
            logger.error(f"PDF conversion failed: output file not created at {output_path}")
            return False
        
        logger.info(f"Successfully converted {excel_path.name} to PDF")
        return True
        
    except ImportError as e:
        logger.error(str(e))
        return False
    except Exception as e:
        logger.error(f"Error converting Excel file {excel_path.name} to PDF: {e}")
        return False
